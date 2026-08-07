#!/usr/bin/env python3
"""Render an .xlsx/.xlsm workbook as markdown so a Read-only blind judge can see its formula layer.

Why this exists
---------------
S13's central discriminator is 活公式 vs 硬编码点值 — whether the five-layer cascade is wired with
live formulas that recompute when an assumption cell changes, or is a table of hand-typed numbers
dressed up as a model. That distinction lives in the formula layer.

The blind rubric judge (`agents/eval-rubric-judge.md`) has `tools: Read` only — deliberately, so it
cannot discover ground truth on disk. That also means it cannot open a workbook. This script is the
bridge: it flattens the workbook into text that shows, per cell, BOTH the formula and its cached
value, so the judge can assess the formula layer without any extra tool.

Two loads are required
----------------------
openpyxl cannot return formulas and cached values from a single load:
  · load_workbook(data_only=False) -> formulas ("=B4*C4"), cached values absent
  · load_workbook(data_only=True)  -> cached values, formulas absent
So we load twice and zip the two views together.

Cached values may be entirely missing
-------------------------------------
A workbook written by openpyxl (or any non-Excel writer) carries no cached results — Excel computes
them on open, and nobody opened it. In that case the data_only view is all None. That is NOT the
same as an empty workbook, and a judge who saw a blank value column would misread it badly. So we
detect the condition and say so, loudly, at the top of the dump.

Usage:
  python workbook_dump.py --in work.xlsx --out work.md [--max-cells-per-sheet 800]

Exit code 0 on success; 2 if the workbook cannot be opened at all.
"""
import argparse
import os
import sys

try:
    import openpyxl
except ImportError:
    print("openpyxl required: pip install openpyxl", file=sys.stderr)
    raise

DEFAULT_MAX_CELLS = 800


def _fmt(v):
    """Render a cell value compactly, without inventing precision."""
    if v is None:
        return ""
    if isinstance(v, float):
        # Trim float noise (0.30000000000000004) without changing the number's meaning.
        s = f"{v:.10g}"
        return s
    return str(v).replace("\n", " ").replace("|", "\\|").strip()


def dump_sheet(ws_f, ws_v, max_cells):
    """Return (lines, stats) for one worksheet.

    ws_f: worksheet from the formulas load; ws_v: same sheet from the cached-values load.
    """
    rows = []
    n_formula = n_const = 0
    truncated = 0

    for row in ws_f.iter_rows():
        for cell in row:
            raw = cell.value
            if raw is None or (isinstance(raw, str) and not raw.strip()):
                continue
            is_formula = isinstance(raw, str) and raw.startswith("=")
            if is_formula:
                n_formula += 1
            else:
                n_const += 1
            if len(rows) >= max_cells:
                truncated += 1
                continue
            try:
                cached = ws_v[cell.coordinate].value
            except Exception:
                cached = None
            rows.append((cell.coordinate,
                         _fmt(raw) if is_formula else "",
                         _fmt(raw) if not is_formula else _fmt(cached)))

    stats = {"formula": n_formula, "const": n_const, "truncated": truncated}
    return rows, stats


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="src", required=True, help="path to the .xlsx/.xlsm workbook")
    ap.add_argument("--out", required=True, help="markdown output path")
    ap.add_argument("--max-cells-per-sheet", type=int, default=DEFAULT_MAX_CELLS,
                    help=f"cap rendered cells per sheet (default {DEFAULT_MAX_CELLS}); "
                         f"any overflow is reported, never silently dropped")
    a = ap.parse_args()

    try:
        wb_f = openpyxl.load_workbook(a.src, data_only=False)
        wb_v = openpyxl.load_workbook(a.src, data_only=True)
    except Exception as e:
        print(f"workbook_dump: cannot open {a.src}: {e}", file=sys.stderr)
        sys.exit(2)

    sheets = []
    total_formula = total_const = total_truncated = 0
    cached_seen = False

    for name in wb_f.sheetnames:
        ws_f = wb_f[name]
        ws_v = wb_v[name] if name in wb_v.sheetnames else ws_f
        rows, stats = dump_sheet(ws_f, ws_v, a.max_cells_per_sheet)
        total_formula += stats["formula"]
        total_const += stats["const"]
        total_truncated += stats["truncated"]
        for _, formula, value in rows:
            if formula and value != "":
                cached_seen = True
        sheets.append((name, rows, stats))

    out = []
    out.append(f"# 工作簿转录：{os.path.basename(a.src)}")
    out.append("")
    out.append(f"共 {len(sheets)} 张工作表；公式单元格 {total_formula} 个，常量单元格 {total_const} 个。")
    out.append("")
    out.append("本文件由 `workbook_dump.py` 从提交的工作簿逐格转录而成。每张表列出「单元格 / 公式 / 值」——")
    out.append("**公式列非空即表示该格是活公式**（改动其引用的假设格会使其重算）；公式列为空则该格是硬编码的常量。")
    out.append("")

    if total_formula == 0:
        out.append("> ⚠ **本工作簿不含任何公式** —— 全部单元格都是硬编码的常量值。")
        out.append("")
    elif not cached_seen:
        out.append("> ⚠ **本工作簿没有缓存值** —— 公式本身完整存在，但从未被 Excel 打开计算过")
        out.append("> （由 openpyxl 等程序写出的工作簿即是如此）。因此下表中公式格的「值」列为空，")
        out.append("> **这不代表公式算不出结果，也不代表表格是空的**。请依据公式列判断建模质量。")
        out.append("")

    if total_truncated:
        out.append(f"> ⚠ 有 {total_truncated} 个单元格超出每表 {a.max_cells_per_sheet} 格的转录上限未被列出"
                   f"（按行列顺序，超出部分在各表末尾）。")
        out.append("")

    for name, rows, stats in sheets:
        out.append(f"## 工作表：{name}")
        out.append("")
        out.append(f"公式格 {stats['formula']} · 常量格 {stats['const']}"
                   + (f" · 未列出 {stats['truncated']}" if stats["truncated"] else ""))
        out.append("")
        if not rows:
            out.append("_（空表）_")
            out.append("")
            continue
        out.append("| 单元格 | 公式 | 值 |")
        out.append("|---|---|---|")
        for coord, formula, value in rows:
            out.append(f"| {coord} | {formula} | {value} |")
        out.append("")

    os.makedirs(os.path.dirname(os.path.abspath(a.out)) or ".", exist_ok=True)
    with open(a.out, "w", encoding="utf-8") as f:
        f.write("\n".join(out))

    print(f"workbook_dump: {a.src} -> {a.out} "
          f"({len(sheets)} sheets, {total_formula} formula cells, {total_const} constant cells"
          + (f", {total_truncated} cells over the cap" if total_truncated else "") + ")")


if __name__ == "__main__":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass
    main()
